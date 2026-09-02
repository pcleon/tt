#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将集群拓扑表格（Sheet1）转换为实例角色扁平表格（Sheet2 格式）的脚本.

输入示例（Sheet1 宽表格式）：
cluster_name  M          S          TS         LS         OS         vip            lower_case_table_names
db1           127.0.0.1  127.0.0.2  127.0.0.3  None       None       127.0.0.200    0

输出示例（Sheet2 扁平格式）：
idc  cluster_name  IP         instance_role
p0   db1           127.0.0.1  M
p0   db1           127.0.0.2  S
p0   db1           127.0.0.3  TS
"""

import argparse
from pathlib import Path
import sys
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# IP 与机房（IDC）映射表（通过 IP 地址第 2 段进行匹配）
IDC_MAP: Dict[str, str] = {
    "192": "p1",
    "196": "p2",
    "200": "p3",
    "999": "p4",
    "0": "p0",
}

# 定义标准角色列名映射
ROLE_MAP: Dict[str, str] = {
    "m": "M",
    "s": "S",
    "ts": "TS",
    "ls": "LS",
    "os": "OS",
}

# 兼容旧版角色列名映射
LEGACY_ROLE_MAP: Dict[str, str] = {
    "m": "M",
    "s": "S",
    "l": "TS",
    "t": "LS",
    "y": "OS",
}


def get_idc_by_ip(ip: str) -> str:
    """根据 IP 地址的第 2 段获取对应的 IDC 机房标识.

    例如 '10.192.1.1' -> 第 2 段为 '192' -> 返回 'p1'；'127.0.0.1' -> 返回 'p0'.
    若无法提取或未在映射表中找到，则返回空字符串 ''.

    Args:
        ip: 点分十进制 IP 地址字符串.

    Returns:
        机房标识字符串（如 'p1', 'p2' 等），未匹配则返回空字符串 ''.
    """
    parts = ip.strip().split(".")
    if len(parts) >= 2:
        second_octet = parts[1].strip()
        return IDC_MAP.get(second_octet, "")
    return ""


def load_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Worksheet:
    """按名称查找工作表（不区分大小写），未找到则返回活动工作表.

    Args:
        wb: openpyxl 工作簿对象.
        sheet_name: 目标工作表名称.

    Returns:
        匹配到的 Worksheet 对象.
    """
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return wb[name]
    return wb.active


def extract_roles(ws: Worksheet) -> List[Tuple[str, str, str, str]]:
    """从源工作表中提取 (idc, cluster_name, IP, instance_role) 记录列表.

    Args:
        ws: openpyxl 的源工作表对象.

    Returns:
        四元组列表，每个元素为 (idc, cluster_name, ip, instance_role).
    """
    if ws.max_row < 1 or ws.max_column < 1:
        return []

    cluster_col = 1
    role_cols: List[Tuple[int, str]] = []
    header_map: Dict[str, int] = {}

    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col_idx).value
        header = str(cell_val or "").strip().lower()
        if not header:
            continue
        if header in ("cluster", "cluster_name", "db", "instance"):
            cluster_col = col_idx
            continue
        header_map[header] = col_idx

    # 判断是否包含新版角色（ts, ls, os）
    if any(k in header_map for k in ("ts", "ls", "os")):
        for raw_key, target_role in ROLE_MAP.items():
            if raw_key in header_map:
                role_cols.append((header_map[raw_key], target_role))
    elif any(k in header_map for k in ("l", "t", "y")):
        for raw_key, target_role in LEGACY_ROLE_MAP.items():
            if raw_key in header_map:
                role_cols.append((header_map[raw_key], target_role))
    else:
        for raw_key, target_role in ROLE_MAP.items():
            if raw_key in header_map:
                role_cols.append((header_map[raw_key], target_role))

    records: List[Tuple[str, str, str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        cluster_val = ws.cell(row=row_idx, column=cluster_col).value
        if cluster_val is None or not str(cluster_val).strip():
            continue
        cluster_name = str(cluster_val).strip()

        for col_idx, role_name in role_cols:
            ip_val = ws.cell(row=row_idx, column=col_idx).value
            if ip_val is None:
                continue
            ip_str = str(ip_val).strip()
            if not ip_str or ip_str in ("-", "none", "null"):
                continue
            idc = get_idc_by_ip(ip_str)
            records.append((idc, cluster_name, ip_str, role_name))

    return records


def print_table(records: List[Tuple[str, str, str, str]]) -> None:
    """在终端格式化打印记录表格.

    Args:
        records: (idc, cluster_name, ip, instance_role) 记录列表.
    """
    headers = ("idc", "cluster_name", "IP", "instance_role")
    col_widths = [len(h) for h in headers]

    for row in records:
        for i, val in enumerate(row):
            col_widths[i] = max(col_widths[i], len(val))

    header_line = "  ".join(h.ljust(col_widths[i]) for i, h in enumerate(headers))
    sep_line = "  ".join("-" * col_widths[i] for i in range(len(headers)))
    print(header_line)
    print(sep_line)

    for row in records:
        print("  ".join(val.ljust(col_widths[i]) for i, val in enumerate(row)))


def generate_sql(records: List[Tuple[str, str, str, str]]) -> str:
    """根据记录列表生成各实例的独立 INSERT SQL 语句.

    每条记录生成一条单独的 insert 语句，格式示例如下：
    insert into table_name (idc,cluster_name,instance_name,ip,port,instance_role) values
    ('p0','db1','127.0.0.2_3306','127.0.0.2',3306,'S');

    Args:
        records: (idc, cluster_name, ip, instance_role) 记录列表.

    Returns:
        包含所有 INSERT 语句的字符串（换行分隔）.
    """
    sql_statements: List[str] = []
    for idc, cluster_name, ip, instance_role in records:
        instance_name = f"{ip}_3306"
        stmt = (
            "insert into table_name (idc,cluster_name,instance_name,ip,port,instance_role) values\n"
            f"('{idc}','{cluster_name}','{instance_name}','{ip}',3306,'{instance_role}');"
        )
        sql_statements.append(stmt)
    return "\n".join(sql_statements)


def export_to_excel(
    records: List[Tuple[str, str, str, str]],
    output_path: Path,
    sheet_name: str = "Sheet2",
    template_path: Optional[Path] = None,
) -> None:
    """将记录导出至 Excel 文件的指定工作表中.

    Args:
        records: (idc, cluster_name, ip, instance_role) 记录列表.
        output_path: 输出 Excel 文件路径.
        sheet_name: 目标工作表名称，默认 'Sheet2'.
        template_path: 模板 Excel 文件路径，若指定且存在则基于该文件修改.
    """
    if template_path and template_path.exists() and template_path.stat().st_size > 0:
        wb = openpyxl.load_workbook(template_path)
    elif output_path.exists() and output_path.stat().st_size > 0:
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()

    target_ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            target_ws = wb[name]
            break

    if target_ws is None:
        target_ws = wb.create_sheet(title=sheet_name)
    else:
        # 清空原工作表已有单元格
        for row in list(target_ws.iter_rows()):
            for cell in row:
                cell.value = None

    target_ws.append(["idc", "cluster_name", "IP", "instance_role"])
    for row in records:
        target_ws.append(list(row))

    wb.save(output_path)
    print(f"\n已成功导出 {len(records)} 条数据至: {output_path} (工作表: {target_ws.title})")


def main() -> None:
    """命令行入口函数，负责解析参数并执行转换与输出."""
    parser = argparse.ArgumentParser(description="将 Sheet1 宽表数据转换为 Sheet2 扁平角色数据或 SQL")
    parser.add_argument("-i", "--input", default="list.xlsx", help="输入的 Excel 文件路径，默认 list.xlsx")
    parser.add_argument("-s", "--sheet", default="Sheet1", help="源工作表名称，默认 Sheet1")
    parser.add_argument("-o", "--output", help="导出的目标文件路径（支持 .xlsx 或 .sql）")
    parser.add_argument("--out-sheet", default="Sheet2", help="导出时的目标工作表名称，默认 Sheet2")
    parser.add_argument("-q", "--sql", action="store_true", help="生成并输出 INSERT SQL 语句")
    args = parser.parse_args()

    src_path = Path(args.input)
    if not src_path.exists():
        print(f"输入文件不存在: {src_path}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = load_sheet(wb, args.sheet)
    records = extract_roles(ws)

    # 若指定了 -q 参数，则生成并输出 SQL
    if args.sql:
        sql_text = generate_sql(records)
        print(sql_text)
        if args.output:
            Path(args.output).write_text(sql_text + "\n", encoding="utf-8")
            print(f"\n已成功将 SQL 保存至: {args.output}")
        return

    # 默认：终端格式化打印
    print_table(records)

    # 若指定了输出目标则导出到 Excel
    if args.output:
        out_path = Path(args.output)
        template_file = src_path if out_path.resolve() == src_path.resolve() else None
        export_to_excel(records, out_path, sheet_name=args.out_sheet, template_path=template_file)


if __name__ == "__main__":
    main()
