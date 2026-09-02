#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从 Excel 表格生成数据库复制拓扑配置文件的脚本."""

import argparse
from pathlib import Path
import sys
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# 定义标准新版角色列名顺序（M, S, TS, LS, OS）
NEW_ROLE_COLUMNS: List[str] = ["m", "s", "ts", "ls", "os"]
# 定义兼容旧版角色列名顺序（m, s, l, t, y）
LEGACY_ROLE_COLUMNS: List[str] = ["m", "s", "l", "t", "y"]


def identify_columns(ws: Worksheet) -> Tuple[int, List[int], Optional[int]]:
    """根据首行表头自动识别集群名列、角色 IP 列（M/S/TS/LS/OS）与 lower_case 参数列.

    优先按 M, S, TS, LS, OS 的顺序匹配列；若未出现 TS/LS/OS 但包含旧版 l/t/y，
    则按旧版 m, s, l, t, y 顺序识别；若未显式匹配到角色列，则回退为自动提取
    排除 VIP 和参数列后的所有有效数据列。

    Args:
        ws: openpyxl 的工作表对象（Worksheet）.

    Returns:
        包含三个元素的元组：
            - db_col: 集群名所在列号（1-based 索引）.
            - ip_cols: 按 M, S, TS, LS, OS 角色顺序排序的节点 IP 列号列表.
            - lower_case_col: lower_case_table_names 参数列号，未找到则为 None.
    """
    db_col = 1
    lower_case_col: Optional[int] = None
    header_to_col = {}
    fallback_ip_cols: List[int] = []

    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col_idx).value
        header = str(cell_val or "").strip().lower()

        if not header:
            continue
        if "vip" in header:
            continue
        if "lower_case" in header:
            lower_case_col = col_idx
            continue

        if col_idx == 1 or header in ("cluster", "db", "instance"):
            db_col = col_idx
            continue

        header_to_col[header] = col_idx
        fallback_ip_cols.append(col_idx)

    # 1. 优先检查是否匹配到新版角色列（含 ts, ls, os）
    if any(k in header_to_col for k in ("ts", "ls", "os")):
        ip_cols = [header_to_col[r] for r in NEW_ROLE_COLUMNS if r in header_to_col]
        return db_col, ip_cols, lower_case_col

    # 2. 检查是否匹配到旧版角色列（含 l, t, y）
    if any(k in header_to_col for k in ("l", "t", "y")):
        ip_cols = [header_to_col[r] for r in LEGACY_ROLE_COLUMNS if r in header_to_col]
        return db_col, ip_cols, lower_case_col

    # 3. 检查仅有 m, s 的情况
    if any(k in header_to_col for k in ("m", "s")):
        ip_cols = [header_to_col[r] for r in ("m", "s") if r in header_to_col]
        return db_col, ip_cols, lower_case_col

    # 4. 若未匹配到特定角色表头，回退使用顺序提取的所有候选 IP 列
    return db_col, fallback_ip_cols, lower_case_col


def format_lower_case_val(raw_val: Optional[object]) -> str:
    """格式化 lower_case_table_names 参数值，空值默认返回 '1'.

    Args:
        raw_val: 单元格原始值.

    Returns:
        格式化后的参数字符串（如 '0' 或 '1'）.
    """
    if raw_val is None or str(raw_val).strip() == "":
        return "1"
    try:
        return str(int(float(raw_val)))
    except (ValueError, TypeError):
        return str(raw_val).strip()


def process_sheet(ws: Worksheet) -> List[str]:
    """读取工作表并生成各集群的配置文本块.

    按 M -> S -> TS -> LS -> OS 顺序线性链式级联，每个有效节点以前一个有效节点作为 rep_master_ip.

    Args:
        ws: openpyxl 的工作表对象.

    Returns:
        按集群组织的配置文本块列表（每个块包含集群名和各节点配置行）.
    """
    db_col, ip_cols, lower_case_col = identify_columns(ws)
    blocks: List[str] = []

    for row_idx in range(2, ws.max_row + 1):
        db_cell = ws.cell(row=row_idx, column=db_col).value
        if not db_cell or not str(db_cell).strip():
            continue
        db_name = str(db_cell).strip()

        # 提取当前行的有效 IP 列表（按角色顺序，已过滤空单元格）
        ips: List[str] = []
        for c in ip_cols:
            val = ws.cell(row=row_idx, column=c).value
            if val is not None and str(val).strip():
                ips.append(str(val).strip())

        if not ips:
            continue

        # 获取参数列取值
        raw_lc = ws.cell(row=row_idx, column=lower_case_col).value if lower_case_col else None
        lc_val = format_lower_case_val(raw_lc)

        # 组装当前集群输出行：按 M -> S -> TS -> LS -> OS 链式级联复制
        lines: List[str] = [f"#{db_name}"]
        for i, ip in enumerate(ips):
            param = f"lower_case_table_names={lc_val}"
            if i == 0:
                lines.append(f"{ip} {param}")
            else:
                lines.append(f"{ip} rep_master_ip={ips[i - 1]}  {param}")

        blocks.append("\n".join(lines))

    return blocks


def load_target_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Worksheet:
    """根据名称查找目标工作表（忽略大小写），若未找到则回退至活动工作表.

    Args:
        wb: openpyxl 工作簿对象.
        sheet_name: 期望读取的工作表名.

    Returns:
        匹配到的 Worksheet 对象.
    """
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return wb[name]
    return wb.active


def main() -> None:
    """命令行入口函数，负责解析参数并执行生成流程."""
    parser = argparse.ArgumentParser(description="从 Excel 表格生成 rep_master_ip 与参数配置")
    parser.add_argument("-i", "--input", default="list.xlsx", help="输入的 Excel 文件路径，默认 list.xlsx")
    parser.add_argument("-s", "--sheet", default="sheet1", help="目标工作表名，默认 sheet1")
    parser.add_argument("-o", "--output", help="输出文件路径；若未指定则输出至标准输出")
    args = parser.parse_args()

    src_path = Path(args.input)
    if not src_path.exists():
        print(f"输入文件不存在: {src_path}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = load_target_sheet(wb, args.sheet)
    blocks = process_sheet(ws)

    output_text = "\n\n".join(blocks) + ("\n" if blocks else "")

    if args.output:
        Path(args.output).write_text(output_text, encoding="utf-8")
    else:
        sys.stdout.write(output_text)


if __name__ == "__main__":
    main()
